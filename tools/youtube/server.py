from __future__ import annotations

import io
import json
import math
import re
import threading
import uuid
import urllib.parse
import urllib.request
import zipfile
from datetime import datetime, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

import yt_dlp


ROOT = Path(__file__).resolve().parent
STATIC = ROOT / "static"
API_BASE = "https://www.googleapis.com/youtube/v3"
MAX_RESULTS = 50
JOBS: dict[str, dict[str, Any]] = {}
JOBS_LOCK = threading.Lock()


class YoutubeApiError(RuntimeError):
    pass


def api_get(path: str, api_key: str, **params: str) -> dict[str, Any]:
    query = {"key": api_key, **params}
    url = f"{API_BASE}/{path}?{urllib.parse.urlencode(query)}"
    request = urllib.request.Request(url, headers={"Accept": "application/json"})

    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        try:
            payload = json.loads(details)
            message = payload.get("error", {}).get("message", details)
        except json.JSONDecodeError:
            message = details or exc.reason
        raise YoutubeApiError(message) from exc
    except urllib.error.URLError as exc:
        raise YoutubeApiError(f"Could not reach YouTube API: {exc.reason}") from exc


def parse_channel_url(raw_url: str) -> dict[str, str]:
    raw_url = raw_url.strip()
    if not raw_url:
        raise ValueError("Paste a YouTube channel link first.")

    if raw_url.startswith("@"):
        return {"kind": "handle", "value": raw_url[1:]}

    parsed = urllib.parse.urlparse(raw_url if "://" in raw_url else f"https://{raw_url}")
    host = parsed.netloc.lower().replace("www.", "")
    path_parts = [urllib.parse.unquote(part) for part in parsed.path.split("/") if part]

    if "youtube.com" not in host and "youtu.be" not in host:
        raise ValueError("Please paste a youtube.com channel URL.")

    if not path_parts:
        raise ValueError("The URL does not include a channel path.")

    first = path_parts[0]
    if first.startswith("@"):
        return {"kind": "handle", "value": first[1:]}

    if first == "channel" and len(path_parts) >= 2:
        return {"kind": "id", "value": path_parts[1]}

    if first == "user" and len(path_parts) >= 2:
        return {"kind": "username", "value": path_parts[1]}

    if first in {"c", "shorts", "videos", "featured", "live", "streams"} and len(path_parts) >= 2:
        return {"kind": "search", "value": path_parts[1]}

    return {"kind": "search", "value": first}


def resolve_channel(channel_url: str, api_key: str) -> dict[str, Any]:
    parsed = parse_channel_url(channel_url)

    if parsed["kind"] == "id":
        data = api_get(
            "channels",
            api_key,
            part="snippet,statistics,contentDetails,brandingSettings",
            id=parsed["value"],
        )
    elif parsed["kind"] == "handle":
        data = api_get(
            "channels",
            api_key,
            part="snippet,statistics,contentDetails,brandingSettings",
            forHandle=parsed["value"],
        )
    elif parsed["kind"] == "username":
        data = api_get(
            "channels",
            api_key,
            part="snippet,statistics,contentDetails,brandingSettings",
            forUsername=parsed["value"],
        )
    else:
        search = api_get(
            "search",
            api_key,
            part="snippet",
            q=parsed["value"],
            type="channel",
            maxResults="1",
        )
        items = search.get("items", [])
        if not items:
            raise YoutubeApiError("Could not find a matching channel for this URL.")
        channel_id = items[0]["snippet"]["channelId"]
        data = api_get(
            "channels",
            api_key,
            part="snippet,statistics,contentDetails,brandingSettings",
            id=channel_id,
        )

    items = data.get("items", [])
    if not items:
        raise YoutubeApiError("No channel was found for this link.")
    return items[0]


def iso_duration_to_seconds(duration: str) -> int:
    pattern = re.compile(
        r"^P(?:(?P<days>\d+)D)?(?:T(?:(?P<hours>\d+)H)?(?:(?P<minutes>\d+)M)?(?:(?P<seconds>\d+)S)?)?$"
    )
    match = pattern.match(duration or "")
    if not match:
        return 0
    parts = {key: int(value or 0) for key, value in match.groupdict().items()}
    return parts["days"] * 86400 + parts["hours"] * 3600 + parts["minutes"] * 60 + parts["seconds"]


def format_duration(seconds: int) -> str:
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60
    if hours:
        return f"{hours}:{minutes:02d}:{secs:02d}"
    return f"{minutes}:{secs:02d}"


def fetch_upload_video_ids(uploads_playlist_id: str, api_key: str, limit: int | None = None) -> list[str]:
    ids: list[str] = []
    page_token = ""

    while True:
        data = api_get(
            "playlistItems",
            api_key,
            part="contentDetails",
            playlistId=uploads_playlist_id,
            maxResults=str(MAX_RESULTS),
            pageToken=page_token,
        )
        for item in data.get("items", []):
            video_id = item.get("contentDetails", {}).get("videoId")
            if video_id:
                ids.append(video_id)
                if limit and len(ids) >= limit:
                    return ids

        page_token = data.get("nextPageToken", "")
        if not page_token:
            return ids


def fetch_video_details(video_ids: list[str], api_key: str) -> list[dict[str, Any]]:
    videos: list[dict[str, Any]] = []
    for offset in range(0, len(video_ids), MAX_RESULTS):
        batch = video_ids[offset : offset + MAX_RESULTS]
        data = api_get(
            "videos",
            api_key,
            part="snippet,contentDetails,statistics,status",
            id=",".join(batch),
            maxResults=str(MAX_RESULTS),
        )
        videos.extend(data.get("items", []))
    return videos


def int_or_blank(value: Any) -> int | str:
    try:
        return int(value)
    except (TypeError, ValueError):
        return ""


def ytdlp_date(upload_date: Any, timestamp: Any) -> tuple[str, str]:
    text = str(upload_date or "")
    if len(text) == 8 and text.isdigit():
        published_date = f"{text[:4]}-{text[4:6]}-{text[6:]}"
        return published_date, f"{published_date}T00:00:00Z"
    if timestamp:
        dt = datetime.fromtimestamp(int(timestamp), timezone.utc)
        return dt.date().isoformat(), dt.isoformat(timespec="seconds")
    return "", ""

def first_text(*values: Any) -> str:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def parse_joined_date(text: str) -> str:
    cleaned = re.sub(r"^\s*Joined\s+", "", text or "", flags=re.IGNORECASE).strip()
    cleaned = cleaned.replace(",", "")
    for fmt in ("%b %d %Y", "%B %d %Y", "%d %b %Y", "%d %B %Y", "%b %Y", "%B %Y", "%Y"):
        try:
            parsed = datetime.strptime(cleaned, fmt)
            return parsed.date().isoformat()
        except ValueError:
            pass
    return text.strip()


def text_from_runs(value: Any) -> str:
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        if value.get("simpleText"):
            return str(value["simpleText"])
        if isinstance(value.get("runs"), list):
            return "".join(str(run.get("text", "")) for run in value["runs"] if isinstance(run, dict))
    return ""


def walk_json(value: Any):
    yield value
    if isinstance(value, dict):
        for child in value.values():
            yield from walk_json(child)
    elif isinstance(value, list):
        for child in value:
            yield from walk_json(child)


def find_channel_about_details(data: Any) -> dict[str, str]:
    details = {"Published At": "", "Country": ""}
    for item in walk_json(data):
        if not isinstance(item, dict):
            continue

        joined = text_from_runs(item.get("joinedDateText"))
        if joined and not details["Published At"]:
            details["Published At"] = parse_joined_date(joined)

        label = text_from_runs(item.get("title") or item.get("label") or item.get("subtitle"))
        value = text_from_runs(item.get("content") or item.get("value") or item.get("text"))
        if label.lower() == "country" and value and not details["Country"]:
            details["Country"] = value

        if item.get("country") and not details["Country"]:
            details["Country"] = text_from_runs(item.get("country"))

    return details


def extract_yt_initial_data(html: str) -> dict[str, Any]:
    marker = "var ytInitialData = "
    start = html.find(marker)
    if start < 0:
        marker = "ytInitialData = "
        start = html.find(marker)
    if start < 0:
        return {}

    start += len(marker)
    end = html.find(";</script>", start)
    if end < 0:
        end = html.find(";</", start)
    if end < 0:
        return {}

    try:
        return json.loads(html[start:end])
    except json.JSONDecodeError:
        return {}


def extract_channel_about_no_api(channel_url: str) -> dict[str, str]:
    details = {"Published At": "", "Country": ""}
    opts: dict[str, Any] = {
        "quiet": True,
        "no_warnings": True,
        "skip_download": True,
        "ignoreerrors": True,
        "extract_flat": "in_playlist",
        "playlistend": 1,
    }
    try:
        with yt_dlp.YoutubeDL(opts) as ydl:
            info = ydl.extract_info(channel_tab_url(channel_url, "about"), download=False) or {}
        details["Published At"] = first_text(info.get("channel_created"), info.get("release_date"))
        details["Country"] = first_text(info.get("country"), info.get("channel_country"))
    except Exception:
        pass

    try:
        request = urllib.request.Request(
            channel_tab_url(channel_url, "about"),
            headers={"User-Agent": "Mozilla/5.0", "Accept-Language": "en-US,en;q=0.9"},
        )
        with urllib.request.urlopen(request, timeout=30) as response:
            html = response.read().decode("utf-8", errors="replace")
        page_details = find_channel_about_details(extract_yt_initial_data(html))
        details["Published At"] = first_text(details["Published At"], page_details["Published At"])
        details["Country"] = first_text(details["Country"], page_details["Country"])
    except Exception:
        pass

    return details

def channel_tab_url(raw_url: str, tab: str) -> str:
    raw_url = raw_url.strip()
    if raw_url.startswith("@"):
        return f"https://www.youtube.com/{raw_url}/{tab}"

    parsed = urllib.parse.urlparse(raw_url if "://" in raw_url else f"https://{raw_url}")
    host = parsed.netloc.lower().replace("www.", "")
    parts = [urllib.parse.unquote(part) for part in parsed.path.split("/") if part]
    if "youtube.com" not in host or not parts:
        raise ValueError("Please paste a YouTube channel URL.")

    first = parts[0]
    if first.startswith("@"):
        base = first
    elif first in {"channel", "user", "c"} and len(parts) >= 2:
        base = f"{first}/{parts[1]}"
    else:
        base = first
    return f"https://www.youtube.com/{base}/{tab}"


def best_thumbnail(info: dict[str, Any]) -> str:
    thumbnails = info.get("thumbnails") or []
    if thumbnails:
        return (thumbnails[-1] or {}).get("url", "")
    return info.get("thumbnail", "") or ""


def preview_channel_no_api(channel_url: str) -> dict[str, Any]:
    opts: dict[str, Any] = {
        "quiet": True,
        "no_warnings": True,
        "skip_download": True,
        "ignoreerrors": True,
        "extract_flat": "in_playlist",
        "playlistend": 1,
    }
    with yt_dlp.YoutubeDL(opts) as ydl:
        info = ydl.extract_info(channel_tab_url(channel_url, "videos"), download=False) or {}
    title = info.get("channel") or info.get("uploader") or info.get("title", "").replace(" - Videos", "")
    return {
        "title": title or "YouTube channel",
        "handle": info.get("uploader_id", ""),
        "channelId": info.get("channel_id") or info.get("id", ""),
        "url": info.get("channel_url") or info.get("uploader_url") or channel_url,
        "description": info.get("description", ""),
        "thumbnail": best_thumbnail(info),
    }


def extract_tab_no_api(channel_url: str, tab: str, limit: int | None, flat: bool = False) -> dict[str, Any]:
    opts: dict[str, Any] = {
        "quiet": True,
        "no_warnings": True,
        "skip_download": True,
        "ignoreerrors": True,
    }
    if flat:
        opts["extract_flat"] = "in_playlist"
    if limit:
        opts["playlistend"] = limit
    with yt_dlp.YoutubeDL(opts) as ydl:
        return ydl.extract_info(channel_tab_url(channel_url, tab), download=False) or {}


def extract_video_no_api(video_id: str) -> dict[str, Any]:
    opts: dict[str, Any] = {
        "quiet": True,
        "no_warnings": True,
        "skip_download": True,
        "ignoreerrors": True,
    }
    with yt_dlp.YoutubeDL(opts) as ydl:
        return ydl.extract_info(f"https://www.youtube.com/watch?v={video_id}", download=False) or {}


def row_from_ytdlp_entry(entry: dict[str, Any], content_type: str) -> dict[str, Any] | None:
    video_id = entry.get("id") or ""
    if not video_id:
        return None
    published_date, published_at = ytdlp_date(entry.get("upload_date"), entry.get("timestamp"))
    seconds = int(entry.get("duration") or 0)
    content_type = "Short" if seconds and seconds <= 60 else "Video"
    link = f"https://www.youtube.com/shorts/{video_id}" if content_type == "Short" else f"https://www.youtube.com/watch?v={video_id}"
    return {
        "Content Type": content_type,
        "Title": entry.get("title", ""),
        "Link": link,
        "Published Date": published_date,
        "Published Month": published_date[:7],
        "Published At": published_at,
        "Duration": format_duration(seconds),
        "Duration Seconds": seconds,
        "Views": int_or_blank(entry.get("view_count")),
        "Likes": int_or_blank(entry.get("like_count")),
        "Comments": int_or_blank(entry.get("comment_count")),
        "Video ID": video_id,
        "Channel": entry.get("channel", ""),
        "Channel ID": entry.get("channel_id", ""),
        "Uploader": entry.get("uploader", ""),
        "Uploader ID": entry.get("uploader_id", ""),
        "Category": ", ".join(entry.get("categories") or []),
        "Tags": ", ".join(entry.get("tags") or []),
        "Resolution": entry.get("resolution", ""),
        "Width": int_or_blank(entry.get("width")),
        "Height": int_or_blank(entry.get("height")),
        "FPS": int_or_blank(entry.get("fps")),
        "Age Limit": int_or_blank(entry.get("age_limit")),
        "Live Status": entry.get("live_status", ""),
        "Playable In Embed": entry.get("playable_in_embed", ""),
        "File Size Approx Bytes": int_or_blank(entry.get("filesize_approx") or entry.get("filesize")),
        "Privacy Status": entry.get("availability", "public") or "public",
        "Description": entry.get("description", ""),
        "Thumbnail": entry.get("thumbnail", ""),
    }


def collect_channel_export_no_api(channel_url: str, limit: int | None = None, progress: Any = None) -> dict[str, Any]:
    if progress:
        progress(5, "Reading videos list")
    videos_info = extract_tab_no_api(channel_url, "videos", limit, flat=True)
    if progress:
        progress(12, "Reading shorts list")
    shorts_info = extract_tab_no_api(channel_url, "shorts", limit, flat=True)

    tasks: list[tuple[str, dict[str, Any]]] = []
    for tab, info in {"videos": videos_info, "shorts": shorts_info}.items():
        content_type = "Short" if tab == "shorts" else "Video"
        for entry in info.get("entries") or []:
            if entry and entry.get("id"):
                tasks.append((content_type, entry))

    total = len(tasks)
    rows_by_id: dict[str, dict[str, Any]] = {}
    channel_source = videos_info or shorts_info

    for index, (content_type, flat_entry) in enumerate(tasks, start=1):
        video_id = flat_entry.get("id", "")
        base_percent = 15
        fetch_percent = 72
        current_percent = base_percent + round((index - 1) / max(total, 1) * fetch_percent)
        if progress:
            progress(current_percent, f"Fetching {index} of {total}: {flat_entry.get('title', video_id)}")

        entry = extract_video_no_api(video_id)
        if not entry:
            entry = flat_entry
        row = row_from_ytdlp_entry(entry, content_type)
        if row:
            rows_by_id[video_id] = row

        if progress:
            done_percent = base_percent + round(index / max(total, 1) * fetch_percent)
            progress(done_percent, f"Fetched {index} of {total}")

    if progress:
        progress(88, "Preparing Excel data")

    rows = sorted(rows_by_id.values(), key=lambda row: row.get("Published At") or "", reverse=True)
    shorts = sum(1 for row in rows if row.get("Content Type") == "Short")
    regular_videos = sum(1 for row in rows if row.get("Content Type") != "Short")
    total_content_views = sum(row["Views"] for row in rows if isinstance(row.get("Views"), int))
    total_content_likes = sum(row["Likes"] for row in rows if isinstance(row.get("Likes"), int))
    total_content_comments = sum(row["Comments"] for row in rows if isinstance(row.get("Comments"), int))
    published_dates = [row["Published Date"] for row in rows if row.get("Published Date")]
    channel_title = channel_source.get("channel") or channel_source.get("uploader") or channel_source.get("title", "").replace(" - Videos", "").replace(" - Shorts", "")
    channel_info_url = channel_source.get("channel_url") or channel_source.get("uploader_url") or channel_url
    about_details = extract_channel_about_no_api(channel_info_url)

    channel_info = {
        "Channel ID": channel_source.get("channel_id") or channel_source.get("id", ""),
        "Title": channel_title,
        "Custom URL": str(channel_source.get("uploader_id", "")).lower(),
        "Channel URL": channel_info_url,
        "Description": channel_source.get("description", ""),
        "Country": first_text(channel_source.get("country"), channel_source.get("channel_country"), about_details["Country"]),
        "Published At": first_text(channel_source.get("channel_created"), channel_source.get("release_date"), about_details["Published At"]),
        "Subscribers": int_or_blank(channel_source.get("channel_follower_count")) or "",
        "Hidden Subscriber Count": "",
        "Total Views": total_content_views,
        "API Video Count": len(rows),
        "Fetched Content Count": len(rows),
        "Number of Videos Published": regular_videos,
        "Number of Shorts Published": shorts,
        "Fetched Content Views": total_content_views,
        "Fetched Content Likes": total_content_likes,
        "Fetched Content Comments": total_content_comments,
        "Latest Published Content Date": max(published_dates) if published_dates else "",
        "First Published Content Date": min(published_dates) if published_dates else "",
        "Uploads Playlist ID": f"UU{str(channel_source.get('channel_id') or channel_source.get('id', ''))[2:]}",
        "Exported At": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }
    return {"channel": channel_info, "content": rows}


def collect_channel_export(channel_url: str, api_key: str, limit: int | None = None) -> dict[str, Any]:
    channel = resolve_channel(channel_url, api_key)
    snippet = channel.get("snippet", {})
    stats = channel.get("statistics", {})
    content_details = channel.get("contentDetails", {})
    uploads_playlist_id = content_details.get("relatedPlaylists", {}).get("uploads")
    if not uploads_playlist_id:
        raise YoutubeApiError("This channel does not expose an uploads playlist.")

    video_ids = fetch_upload_video_ids(uploads_playlist_id, api_key, limit)
    videos = fetch_video_details(video_ids, api_key)

    rows: list[dict[str, Any]] = []
    shorts = 0
    regular_videos = 0
    total_content_views = 0
    total_content_likes = 0
    total_content_comments = 0
    published_dates: list[str] = []

    for video in videos:
        video_snippet = video.get("snippet", {})
        video_stats = video.get("statistics", {})
        content = video.get("contentDetails", {})
        seconds = iso_duration_to_seconds(content.get("duration", ""))
        content_type = "Short" if seconds <= 60 else "Video"
        published_at = video_snippet.get("publishedAt", "")
        published_date = published_at[:10]
        published_month = published_at[:7]
        views = int_or_blank(video_stats.get("viewCount"))
        likes = int_or_blank(video_stats.get("likeCount"))
        comments = int_or_blank(video_stats.get("commentCount"))

        if content_type == "Short":
            shorts += 1
        else:
            regular_videos += 1
        if isinstance(views, int):
            total_content_views += views
        if isinstance(likes, int):
            total_content_likes += likes
        if isinstance(comments, int):
            total_content_comments += comments
        if published_date:
            published_dates.append(published_date)

        video_id = video.get("id", "")
        link = (
            f"https://www.youtube.com/shorts/{video_id}"
            if content_type == "Short"
            else f"https://www.youtube.com/watch?v={video_id}"
        )
        rows.append(
            {
                "Content Type": content_type,
                "Title": video_snippet.get("title", ""),
                "Link": link,
                "Published Date": published_date,
                "Published Month": published_month,
                "Published At": published_at,
                "Duration": format_duration(seconds),
                "Duration Seconds": seconds,
                "Views": views,
                "Likes": likes,
                "Comments": comments,
                "Video ID": video_id,
                "Privacy Status": video.get("status", {}).get("privacyStatus", ""),
                "Description": video_snippet.get("description", ""),
                "Thumbnail": video_snippet.get("thumbnails", {}).get("high", {}).get("url", ""),
            }
        )

    channel_info = {
        "Channel ID": channel.get("id", ""),
        "Title": snippet.get("title", ""),
        "Custom URL": snippet.get("customUrl", ""),
        "Channel URL": f"https://www.youtube.com/channel/{channel.get('id', '')}",
        "Description": snippet.get("description", ""),
        "Country": snippet.get("country", ""),
        "Published At": snippet.get("publishedAt", ""),
        "Subscribers": int_or_blank(stats.get("subscriberCount")),
        "Hidden Subscriber Count": stats.get("hiddenSubscriberCount", ""),
        "Total Views": int_or_blank(stats.get("viewCount")),
        "API Video Count": int_or_blank(stats.get("videoCount")),
        "Fetched Content Count": len(rows),
        "Number of Videos Published": regular_videos,
        "Number of Shorts Published": shorts,
        "Fetched Content Views": total_content_views,
        "Fetched Content Likes": total_content_likes,
        "Fetched Content Comments": total_content_comments,
        "Latest Published Content Date": max(published_dates) if published_dates else "",
        "First Published Content Date": min(published_dates) if published_dates else "",
        "Uploads Playlist ID": uploads_playlist_id,
        "Exported At": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }

    return {"channel": channel_info, "content": rows}


def column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def clean_excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    text = str(value)
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)



def excel_image_formula(url: Any) -> str:
    text = clean_excel_value(url)
    if not text:
        return ""
    escaped = str(text).replace('"', '""')
    return f'=IMAGE("{escaped}", "Thumbnail", 0)'

def append_rows(sheet: Any, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append([clean_excel_value(value) for value in row])


def make_xlsx(export: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    content = export["content"]
    channel = export["channel"]
    overview_rows = [
        ["Metric", "Value"],
        ["Channel Name", channel.get("Title", "")],
        ["Channel URL", channel.get("Channel URL", "")],
        ["Subscribers", channel.get("Subscribers", "")],
        ["Total Channel Views", channel.get("Total Views", "")],
        ["Total Content In Channel API", channel.get("API Video Count", "")],
        ["Total Content Fetched", channel.get("Fetched Content Count", "")],
        ["Number of Videos Published", channel.get("Number of Videos Published", "")],
        ["Number of Shorts Published", channel.get("Number of Shorts Published", "")],
        ["Fetched Content Views", channel.get("Fetched Content Views", "")],
        ["Fetched Content Likes", channel.get("Fetched Content Likes", "")],
        ["Fetched Content Comments", channel.get("Fetched Content Comments", "")],
        ["Latest Published Content Date", channel.get("Latest Published Content Date", "")],
        ["First Published Content Date", channel.get("First Published Content Date", "")],
        ["Channel Created At", channel.get("Published At", "")],
        ["Country", channel.get("Country", "")],
        ["Channel ID", channel.get("Channel ID", "")],
        ["Custom URL", channel.get("Custom URL", "")],
        ["Uploads Playlist ID", channel.get("Uploads Playlist ID", "")],
        ["Exported At", channel.get("Exported At", "")],
        [],
        ["Channel Description", channel.get("Description", "")],
    ]

    headers = [
        "Content Type",
        "Title",
        "Link",
        "Published Date",
        "Duration",
        "Views",
        "Likes",
        "Comments",
        "Video ID",
        "Description",
        "Thumbnail",
    ]
    content_rows = [headers]
    for row in content:
        content_rows.append([excel_image_formula(row.get(header, "")) if header == "Thumbnail" else row.get(header, "") for header in headers])

    month_totals: dict[str, dict[str, int]] = {}
    for row in content:
        month = str(row.get("Published Month") or "Unknown")
        bucket = month_totals.setdefault(month, {"Videos": 0, "Shorts": 0, "Total": 0})
        if row.get("Content Type") == "Short":
            bucket["Shorts"] += 1
        else:
            bucket["Videos"] += 1
        bucket["Total"] += 1
    monthly_rows = [["Published Month", "Videos", "Shorts", "Total"]]
    for month in sorted(month_totals.keys(), reverse=True):
        bucket = month_totals[month]
        monthly_rows.append([month, bucket["Videos"], bucket["Shorts"], bucket["Total"]])

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    all_content = workbook.create_sheet("All Content")
    monthly = workbook.create_sheet("Monthly Summary")

    append_rows(overview, overview_rows)
    append_rows(all_content, content_rows)
    append_rows(monthly, monthly_rows)

    for sheet in (overview, all_content, monthly):
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="D92525")
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            width = min(max(len(str(cell.value or "")) for cell in column_cells[:100]) + 2, 55)
            sheet.column_dimensions[letter].width = width

    thumbnail_column = headers.index("Thumbnail") + 1
    all_content.column_dimensions[get_column_letter(thumbnail_column)].width = 24
    for row_index in range(2, all_content.max_row + 1):
        all_content.row_dimensions[row_index].height = 72
    hyperlink_columns = {header: index + 1 for index, header in enumerate(headers) if header in {"Link"}}
    for row in all_content.iter_rows(min_row=2, max_row=all_content.max_row):
        for index in hyperlink_columns.values():
            cell = row[index - 1]
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def set_job(job_id: str, **updates: Any) -> None:
    with JOBS_LOCK:
        job = JOBS.setdefault(job_id, {})
        job.update(updates)


def get_job(job_id: str) -> dict[str, Any] | None:
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        return dict(job) if job else None


def run_export_job(job_id: str, channel_url: str, limit: int | None) -> None:
    try:
        def progress(percent: int, message: str) -> None:
            set_job(job_id, status="running", percent=percent, message=message)

        progress(3, "Starting export")
        export = collect_channel_export_no_api(channel_url, limit, progress)
        progress(92, "Building Excel workbook")
        workbook = make_xlsx(export)
        filename = f"{safe_filename(export['channel'].get('Title') or 'youtube-channel')}.xlsx"
        set_job(
            job_id,
            status="done",
            percent=100,
            message="Excel file is ready",
            filename=filename,
            workbook=workbook,
        )
    except Exception as exc:
        set_job(job_id, status="error", percent=100, message=str(exc), error=str(exc))


class Handler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path in {"/", "/index.html"}:
            self.send_file(STATIC / "index.html", "text/html; charset=utf-8")
            return
        if parsed.path == "/styles.css":
            self.send_file(STATIC / "styles.css", "text/css; charset=utf-8")
            return
        if parsed.path == "/app.js":
            self.send_file(STATIC / "app.js", "application/javascript; charset=utf-8")
            return
        if parsed.path == "/export/status":
            job_id = urllib.parse.parse_qs(parsed.query).get("id", [""])[0]
            job = get_job(job_id)
            if not job:
                self.send_json(404, {"error": "Export job was not found."})
                return
            self.send_json(200, {key: value for key, value in job.items() if key != "workbook"})
            return
        if parsed.path == "/export/download":
            job_id = urllib.parse.parse_qs(parsed.query).get("id", [""])[0]
            job = get_job(job_id)
            if not job or job.get("status") != "done" or not job.get("workbook"):
                self.send_json(404, {"error": "Excel file is not ready yet."})
                return
            workbook = job["workbook"]
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{job.get("filename", "youtube-channel.xlsx")}"')
            self.send_header("Content-Length", str(len(workbook)))
            self.end_headers()
            self.wfile.write(workbook)
            return
        self.send_error(404)

    def do_POST(self) -> None:
        parsed = urllib.parse.urlparse(self.path)

        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            channel_url = payload.get("channelUrl", "").strip()
            limit_text = str(payload.get("limit", "")).strip()
            limit = int(limit_text) if limit_text else None

            if limit is not None and limit <= 0:
                raise ValueError("Limit must be empty or greater than zero.")

            if parsed.path == "/preview":
                self.send_json(200, preview_channel_no_api(channel_url))
                return

            if parsed.path == "/export/start":
                job_id = uuid.uuid4().hex
                set_job(job_id, status="queued", percent=0, message="Waiting to start")
                thread = threading.Thread(target=run_export_job, args=(job_id, channel_url, limit), daemon=True)
                thread.start()
                self.send_json(200, {"jobId": job_id})
                return

            if parsed.path == "/export":
                api_key = payload.get("apiKey", "").strip()
                export = collect_channel_export(channel_url, api_key, limit) if api_key else collect_channel_export_no_api(channel_url, limit)
                workbook = make_xlsx(export)
                filename = safe_filename(export["channel"].get("Title") or "youtube-channel")

                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}.xlsx"')
                self.send_header("Content-Length", str(len(workbook)))
                self.end_headers()
                self.wfile.write(workbook)
                return

            self.send_error(404)
        except (ValueError, YoutubeApiError, json.JSONDecodeError) as exc:
            self.send_json(400, {"error": str(exc)})
        except Exception as exc:
            self.send_json(500, {"error": f"Unexpected error: {exc}"})

    def send_file(self, path: Path, content_type: str) -> None:
        if not path.exists():
            self.send_error(404)
            return
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_json(self, status: int, payload: dict[str, Any]) -> None:
        data = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def safe_filename(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", name).strip("-")
    return cleaned[:80] or "youtube-channel"


def main() -> None:
    server = ThreadingHTTPServer(("127.0.0.1", 8000), Handler)
    print("YouTube Channel Excel Exporter running at http://127.0.0.1:8000")
    server.serve_forever()


if __name__ == "__main__":
    main()

